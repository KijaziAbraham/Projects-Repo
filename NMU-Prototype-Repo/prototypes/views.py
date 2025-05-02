from django.contrib.auth import get_user_model
from django.db.models import Case, When, Value, IntegerField
from rest_framework import viewsets,  filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
import openpyxl
from django.http import HttpResponse
from weasyprint import HTML
from django.template.loader import render_to_string
from .permissions import IsPrototypeOwner, IsAdmin, IsStaff, IsStudent, IsOwnerOrReadOnly, IsReviewer
from .serializers import (
    UserSerializer, PrototypeSerializer, PrototypeAttachmentSerializer, 
    DepartmentSerializer, PrototypeReviewSerializer,  
)
from .models import CustomUser, Prototype, PrototypeAttachment, Department
import logging
from django.db.models import Q
from django.contrib.auth import update_session_auth_hash
logger = logging.getLogger(__name__)
User = get_user_model()
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.parsers import JSONParser
from django.db.models.functions import TruncMonth
from django.db.models import Count
from datetime import datetime
from django.utils import timezone
from datetime import timedelta
from collections import defaultdict
from .serializers import GeneralUserRegistrationSerializer
from rest_framework import generics
from rest_framework.permissions import IsAdminUser
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.validators import RegexValidator
from prototypes.models import CustomUser, Department
import pandas as pd
import logging
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import ValidationError
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework import viewsets, permissions
from .models import PrototypeComment
from .serializers import PrototypeCommentSerializer

class PrototypeCommentViewSet(viewsets.ModelViewSet):
    queryset = PrototypeComment.objects.all()
    serializer_class = PrototypeCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class GeneralUserRegistrationView(generics.CreateAPIView):
    serializer_class = GeneralUserRegistrationSerializer

    def post(self, request, *args, **kwargs):
        data = request.data.copy()
        data['role'] = 'general_user'
        data['is_approved'] = False
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"message": "Registration successful. Awaiting admin approval."}, status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH"])  
@permission_classes([IsAuthenticated])
def user_profile(request):
    """Return or update logged-in user's details"""
    user = request.user

    if request.method == "GET":
        return Response({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": user.role,
            "phone": user.phone,
            "institution_id": user.institution_id,
            "level": user.level,
            "full_name": user.full_name,
            "department": user.department.name if user.department else None,
            "is_approved": user.is_approved,
        })

    elif request.method == "PATCH":
        data = request.data
        user.phone = data.get("phone", user.phone)  
        user.email = data.get("email", user.email)  
        user.save()

        return Response({
            "message": "Profile updated successfully",
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": user.role,
            "phone": user.phone,
            "institution_id": user.institution_id,
            "level": user.level,
        }, status=status.HTTP_200_OK)


class UserCreateAPIView(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class UserViewSet(viewsets.ModelViewSet):
    """View to manage users"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return User.objects.all() if user.role == "admin" else User.objects.filter(id=user.id)

    @action(detail=False, methods=["GET"], permission_classes=[IsAuthenticated])
    def students(self, request):
        """Allow admin, staff, and students to view all student users."""
        
        # Allow all authenticated users to view, but restrict non-admins from modifying
        if request.method != "GET":
            return Response({"error": "You are not allowed to modify this list."}, status=status.HTTP_403_FORBIDDEN)

        students = User.objects.filter(role="student")
        serializer = self.get_serializer(students, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["GET"], permission_classes=[IsAuthenticated])
    def supervisors(self, request):
        """Retrieve all staff members who act as supervisors"""
        supervisors = User.objects.filter(Q(role="staff") | Q(role="admin"))
        serializer = self.get_serializer(supervisors, many=True)
        return Response(serializer.data)

class PrototypeViewSet(viewsets.ModelViewSet):
    """Manage prototypes and provide role-based filtering"""
    queryset = Prototype.objects.all()
    serializer_class = PrototypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'barcode', 'storage_location']
    ordering_fields = ['submission_date']
    parser_classes = (MultiPartParser, FormParser)

    def get_queryset(self):
        """Ensure students see their own prototypes first"""
        user = self.request.user
        queryset = Prototype.objects.all()

        if user.role == "student":
            return queryset.annotate(
                priority=Case(
                    When(student=user, then=Value(0)), 
                    default=Value(1),
                    output_field=IntegerField(),
                )
            ).order_by("priority", "-submission_date")

        elif user.role == "staff":
            return queryset  

        return queryset         # Admin and staff can see all prototypes

    @action(detail=False, methods=['GET'], permission_classes=[IsAuthenticated])
    def all_prototypes(self, request):
        """Return all prototypes for staff & admin."""
        if request.user.role in ['staff', 'admin']:
            prototypes = Prototype.objects.all()
        else:
            return Response({"error": "Unauthorized access."}, status=403)

        serializer = PrototypeSerializer(prototypes, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['POST'], parser_classes=[JSONParser])
    def assign_storage(self, request, pk=None):
        """Allow admins to assign a storage location"""
        user = request.user
        if user.role != 'admin':
            return Response({"error": "Only admins can assign storage locations."}, status=403)

        prototype = self.get_object()
        storage_location = request.data.get("storage_location", "").strip()

        if not prototype.has_physical_prototype:
            return Response({"error": "This prototype does not have a physical version."}, status=400)

        if not storage_location:
            return Response({"error": "Storage location is required."}, status=400)

        prototype.storage_location = storage_location
        prototype.save()

        # Return the updated prototype object
        serializer = PrototypeSerializer(prototype)
        return Response(serializer.data)


    @action(detail=True, methods=['POST'], permission_classes=[IsAuthenticated], parser_classes=[JSONParser])
    def review_prototype(self, request, pk=None):
        """Staff and Admin can review a specific prototype (approval and feedback)."""
        user = request.user

        # Ensure the user is either staff or admin
        if user.role not in ["staff", "admin"]:
            return Response({"error": "Only staff and admins can review prototypes."}, status=status.HTTP_403_FORBIDDEN)

        # Get the prototype object
        try:
            prototype = self.get_object()
        except Prototype.DoesNotExist:
            return Response({"error": "Prototype not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get and validate feedback data from the request
        feedback = request.data.get("feedback", "").strip()

        if not feedback:
            return Response({"error": "Feedback is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the prototype status is updated correctly (if it's not already reviewed)
        if prototype.status == "submitted_reviewed":
            return Response({"error": "Prototype has already been reviewed."}, status=status.HTTP_400_BAD_REQUEST)

        # Update the prototype with feedback and reviewed status
        prototype.status = "submitted_reviewed"
        prototype.feedback = feedback
        prototype.reviewer = user  # Record the staff/admin who reviewed
        prototype.save()

        return Response({"message": "Prototype reviewed and approved successfully."}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=["GET"])
    def storage_locations(self, request):
        """Retrieve all unique storage locations"""
        locations = Prototype.objects.exclude(storage_location__isnull=True).exclude(storage_location="").values_list("storage_location", flat=True).distinct()

        return Response(list(locations))

    @action(detail=False, methods=['GET'])
    def export_excel(self, request):
        """Export prototypes as an Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Title", "Barcode", "Storage Location", "Has Physical Prototype"])

        for proto in Prototype.objects.all():
            ws.append([proto.id, proto.title, proto.barcode, proto.storage_location, proto.has_physical_prototype])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="prototypes.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['GET'])
    def export_pdf(self, request):
        """Export prototypes as a PDF file"""
        prototypes = Prototype.objects.all()
        html_content = render_to_string("export_template.html", {"prototypes": prototypes})
        pdf_file = HTML(string=html_content).write_pdf()

        response = HttpResponse(pdf_file, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="prototypes.pdf"'
        return response


    def has_permission(self, request, view):
        if request.user.role == 'general_user' and view.action not in ['list', 'retrieve']:
            return False
        return super().has_permission(request, view)
    


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post']




@api_view(["POST"])
@permission_classes([IsAuthenticated])
def change_password(request):
    """Allow authenticated users to change their password"""
    user = request.user
    current_password = request.data.get("current_password")
    new_password = request.data.get("new_password")

    if not user.check_password(current_password):
        return Response({"detail": "Current password is incorrect."}, status=400)

    if len(new_password) < 6:
        return Response({"detail": "New password must be at least 6 characters long."}, status=400)

    user.set_password(new_password)
    user.save()

    # Keep the user logged in after password change
    update_session_auth_hash(request, user)

    return Response({"detail": "Password updated successfully!"})



@api_view(['GET'])
def prototype_count_view(request):
    user = request.user
    available_count = Prototype.objects.count()
    
    if user.role == 'student':
        user_count = Prototype.objects.filter(student=user).count()
    else:
        # admin or staff can see all
        user_count = available_count

    return Response({
        'your_count': user_count,
        'available_count': available_count,
    })

@api_view(['GET'])
def upload_summary_30_days(request):
    today = timezone.now()  # aware datetime
    start_date = today - timedelta(days=30)

    # Only use timezone-aware filtering
    prototypes = Prototype.objects.filter(submission_date__gte=start_date)

    # Initialize day counts
    days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    upload_counts = defaultdict(int)

    for prototype in prototypes:
        # Make sure submission_date is timezone-aware
        submission_date = prototype.submission_date
        if timezone.is_naive(submission_date):
            submission_date = timezone.make_aware(submission_date)
        weekday = submission_date.strftime('%a')  # 'Mon', 'Tue', ...
        upload_counts[weekday] += 1

    # Ensure all 7 days are present
    data = [{"day": day, "uploads": upload_counts.get(day, 0)} for day in days_of_week]

    return Response(data)


class AdminUserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        # Optional: Filter out superuser accounts from view
        return CustomUser.objects.exclude(is_superuser=True)

    @action(detail=False, methods=['get'])
    def general_users(self, request):
        general_users = CustomUser.objects.filter(role='general_user')
        serializer = self.get_serializer(general_users, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve_user(self, request, pk=None):
        user = self.get_object()
        if user.role != 'general_user':
            return Response({'detail': 'Only general users can be approved.'}, status=status.HTTP_400_BAD_REQUEST)
        user.is_approved = True
        user.save()
        return Response({'detail': f'User {user.username} approved.'})
    

class UserImportTemplateView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        """Generate and download professional user import template"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "User Import Template"
            
            # Defining styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(
                start_color="64A293",  
                end_color="64A293",
                fill_type="solid"
            )
            description_font = Font(italic=True, color="808080")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_aligned = Alignment(horizontal='center')
            
            # Defining data - Removed username column
            headers = [
                ("Email", "Required (must be unique email address)"),
                ("Full Name", "Required"),
                ("Role", "Required (select from dropdown)"),
                ("Department Code", "Optional (must match existing codes)"),
                ("Phone", "Required"),
                ("Institution ID", "Optional"),
                ("Level", "Required if Role=Student")
            ]
            
            # University header
            ws.merge_cells('A1:G1')
            ws['A1'] = "NM UNIVERSITY - USER IMPORT TEMPLATE"
            ws['A1'].font = Font(bold=True, size=14, color="64A293")
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Instructions
            ws.merge_cells('A3:G5')
            ws['A3'] = """INSTRUCTIONS:
1. Fill in all required fields (Email, Full Name, Phone, Role)
2. For dropdown fields, select from the available options
3. Students must include Level (phd/masters)
4. Department codes must match existing values
5. Password will be set to the user's email initially"""
            ws['A3'].font = Font(size=10)
            ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
            
            # Header row
            for col, (header, _) in enumerate(headers, start=1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_aligned
                # Set column width
                ws.column_dimensions[get_column_letter(col)].width = 22
            
            # Description row
            for col, (_, description) in enumerate(headers, start=1):
                cell = ws.cell(row=8, column=col, value=description)
                cell.font = description_font
                cell.border = thin_border
            
            # Data validation for role field (admin/staff/student only)
            role_dv = DataValidation(
                type="list",
                formula1='"admin,staff,student"',
                allow_blank=False,
                showErrorMessage=True,
                error="Must be admin, staff, or student"
            )
            ws.add_data_validation(role_dv)
            role_dv.add("C9:C1048576")  
            
            # Data validation for level field
            level_dv = DataValidation(
                type="list",
                formula1='"phd,masters"',
                allow_blank=True
            )
            ws.add_data_validation(level_dv)
            level_dv.add("G9:G1048576")  
            
            # Example data row
            example_data = [
                "herieth@nmu.edu",
                "Herieth John",
                "student",  # Dropdown
                "CS",
                "+1234567890",
                "STD12345",
                "masters"  # Dropdown
            ]
            for col, value in enumerate(example_data, start=1):
                cell = ws.cell(row=9, column=col, value=value)
                cell.border = thin_border
            
            # Freeze header row
            ws.freeze_panes = "A9"
            
            # Create binary response
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = (
                'attachment; filename="NMU_User_Import_Template.xlsx"'
            )
            return response
            
        except Exception as e:
            logger.error(f"Template generation error: {str(e)}", exc_info=True)
            return Response(
                {"error": "Failed to generate template"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BulkUserImportView(APIView):
    permission_classes = [IsAdminUser]
    parser_classes = [MultiPartParser]

    @transaction.atomic
    def post(self, request):
        try:
            if 'excel_file' not in request.FILES:
                return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

            excel_file = request.FILES['excel_file']

            try:
                df = pd.read_excel(excel_file, header=6, skiprows=[7])
                df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            except Exception as e:
                return Response({'error': f'Invalid Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

            expected_columns = {
                'email': ['email', 'e-mail', 'email_address'],
                'full_name': ['full_name', 'name', 'fullname'],
                'role': ['role', 'user_role', 'account_type'],
                'phone': ['phone', 'phone_number', 'mobile'],
                'department_code': ['department_code', 'dept_code', 'department'],
                'institution_id': ['institution_id', 'id_number', 'student_id'],
                'level': ['level', 'student_level', 'degree_level']
            }

            column_mapping = {}
            missing_columns = []

            for expected, alternatives in expected_columns.items():
                found = False
                for alt in alternatives:
                    if alt in df.columns:
                        column_mapping[expected] = alt
                        found = True
                        break
                if not found and expected in ['email', 'full_name', 'role', 'phone']:
                    missing_columns.append(expected)

            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            success_count = 0
            errors = []
            valid_roles = ['admin', 'staff', 'student']
            valid_levels = ['phd', 'masters']
            phone_regex = re.compile(r'^\+?\d{7,15}$')

            for index, row in df.iterrows():
                row_num = index + 2
                try:
                    email = row[column_mapping['email']]
                    full_name = row[column_mapping['full_name']]
                    role = row[column_mapping['role']].lower() if pd.notna(row[column_mapping['role']]) else None
                    phone = row[column_mapping['phone']]

                    if pd.isna(email): raise ValidationError('Email is required')
                    if pd.isna(full_name): raise ValidationError('Full name is required')
                    if pd.isna(role): raise ValidationError('Role is required')
                    if pd.isna(phone): raise ValidationError('Phone is required')

                    if role not in valid_roles:
                        raise ValidationError(f'Invalid role: {role}. Must be admin, staff, or student')

                    if not phone_regex.match(str(phone)):
                        raise ValidationError('Invalid phone number format')

                    department = None
                    if 'department_code' in column_mapping and pd.notna(row[column_mapping['department_code']]):
                        try:
                            department = Department.objects.get(code=row[column_mapping['department_code']])
                        except Department.DoesNotExist:
                            raise ValidationError(f'Department not found: {row[column_mapping["department_code"]]}')

                    institution_id = row[column_mapping['institution_id']] if 'institution_id' in column_mapping else ''

                    level = None
                    if role == 'student':
                        if 'level' not in column_mapping or pd.isna(row[column_mapping['level']]):
                            raise ValidationError('Students require level (phd/masters)')
                        level = row[column_mapping['level']].lower()
                        if level not in valid_levels:
                            raise ValidationError(f'Invalid level: {level}. Must be phd or masters')

                    if CustomUser.objects.filter(email=email).exists():
                        raise ValidationError('User with this email already exists')

                    # Generate unique username from full_name
                    base_username = ''.join(full_name.lower().split())
                    username = base_username
                    counter = 1
                    while CustomUser.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    user = CustomUser(
                        username=username,
                        email=email,
                        full_name=full_name,
                        role=role,
                        phone=phone,
                        department=department,
                        institution_id=institution_id,
                        is_approved=True,
                        level=level if role == 'student' else None,
                        is_staff=(role == 'admin')
                    )
                    user.set_password(email)
                    user.save()

                    success_count += 1

                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'error': str(e),
                        'data': {k: v for k, v in row.items() if pd.notna(v)}
                    })

            response_data = {
                'success_count': success_count,
                'error_count': len(errors),
                'errors': errors[:100]
            }

            return Response(response_data, status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Bulk import failed: {str(e)}", exc_info=True)
            return Response({'error': 'Internal server error during import'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PrototypeCommentViewSet(viewsets.ModelViewSet):
    queryset = PrototypeComment.objects.all()
    serializer_class = PrototypeCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
